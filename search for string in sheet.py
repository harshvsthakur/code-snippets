import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\harshvardhans\\Desktop\\python\\test.xlsx', data_only = True )
sheet = wb['Index']

for row in range(1, sheet.max_row + 1):
    for column in "G":  # Here you can add or reduce the columns
        cell_name = "{}{}".format(column, row)
        if sheet[cell_name].value == "L004":
            print (cell_name)
            print (row)
            print (sheet.cell(column = 8, row = row).value[:15])
